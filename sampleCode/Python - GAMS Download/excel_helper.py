import logging
import os.path

from openpyxl import load_workbook

class ExcelHelper:
    """
    This class holds static methods to interact with Excel workbooks
    """

    @staticmethod
    def load_combined_region_builder_regions_sheet(filepath: str) -> dict[str, list[int]] | None:
        """
        This method opens a filled-out Combined Region Builder xlsx (Excel) file and parses the data
        into a dictionary mapping the Region Name to one or more FIPS Codes, which will become
        combined regions
        :param filepath: The path to the .xlsx file to be opened. May be absolute or relative.
        :return: A dictionary where the Keys are the Region names and the Values are the lists of FIPS codes
        to be combined under that Region Name
        """

        # Validate the file actually exists
        if not os.path.exists(filepath):
            logging.error(f"Could not find Combined Region Builder xlsx file at '{filepath}'")
            return None

        # We need to aggregate the contents of the sheet into a dictionary
        # The key will be the 'Region Name' so that all FIPS codes can be clustered together
        region_dict: dict[str, list[int]] = {}

        # Open the workbook from the file path (in read-only mode and ignoring formulas) [for speed]
        workbook = load_workbook(filename=filepath, read_only=True, data_only=True)

        # There is only one sheet, 'Regions'
        sheet = workbook.active

        # Get the last row defined in the xlsx file
        end_row: int = sheet.max_row

        # Process every row after the header
        for row in sheet.iter_rows(min_row=2, max_row=end_row, values_only=True):
            # If the row is blank, skip it
            if all(value is None for value in row):
                continue

            # We must have a Region Name and a FIPS Code, all other cells are ignored
            if row[0] is None or row[2] is None:
                logging.error(f"Sheet Row '{row}' is invalid: Must have Region Name and FIPS Code")
                return None

            region_name: str = row[0]
            fips_code: int = int(row[2])

            # Does this region name already exist in the dictionary?
            if region_dict.get(region_name) is None:
                # If it does not, this is the first FIPS for this region
                region_dict[region_name] = [fips_code]
            else:
                # Otherwise this is another FIPS code for the same region
                region_dict[region_name].append(fips_code)

        return region_dict